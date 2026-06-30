import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

TRUCK_CAPACITIES = {
    '6FT_TRUCK': 1000,
    '8FT_TRUCK': 2000,
    '10FT_TRUCK': 3000,
    '14FT_TRUCK': 4667,
    '17FT_TRUCK': 5467,
    '20FT_TRUCK': 7600,
    '22FT_TRUCK': 8400,
    '24FT_TRUCK': 8934,
    '32FT_TRUCK': 12000,
    '40FT_TRUCK': 16800
}

routes = pd.read_csv('optimized_routes.csv')
sop = pd.read_excel('details.xlsx', sheet_name='S&OP')

sop['DH_clean'] = sop['DH'].astype(str).str.strip()
demand_lookup = dict(zip(sop['DH_clean'], sop['w1']))
restriction_lookup = dict(zip(sop['DH_clean'], sop['Vehicle restrictions']))

def get_node_demand(dh_name):
    clean_name = dh_name.split('_part')[0]
    full_dem = demand_lookup.get(clean_name, 0)
    
    if '_part' not in dh_name:
        return full_dem
        
    res_str = str(restriction_lookup.get(clean_name, '')).strip()
    cap = TRUCK_CAPACITIES.get(res_str, 16800)
    
    part_num = int(dh_name.split('_part')[1])
    remaining = full_dem - (part_num - 1) * cap
    return min(remaining, cap)

def is_split(dh_name):
    return "Yes" if '_part' in dh_name else "No"

master_data = []
max_stops = 0

for _, row in routes.iterrows():
    stops = [s.strip() for s in str(row['Route']).split('->')]
    dhs = [s for s in stops if s != 'BLR-DRY-MH-SUMADHURA' and s != '']
    
    if len(dhs) == 0: continue
    max_stops = max(max_stops, len(dhs))
    
    route_type = "Direct" if len(dhs) == 1 else "Milk Run"
    
    base_dict = {
        'Route Type': route_type,
        'Assigned Vehicle': row['Assigned_Truck'],
        'Vehicle Capacity (kg)': row['Truck_Capacity'],
        'Total Route Demand (kg)': row['Total_Load'],
        'Utilization (%)': row['Utilization_%'],
        'Total Distance (km)': row['Total_Distance_km']
    }
    
    for i, dh in enumerate(dhs):
        clean_name = dh.split('_part')[0]
        dem = get_node_demand(dh)
        restr = str(restriction_lookup.get(clean_name, 'None')).strip()
        split_flag = is_split(dh)
        
        base_dict[f'Stop {i+1} DH'] = clean_name
        base_dict[f'Stop {i+1} Restriction'] = restr
        base_dict[f'Stop {i+1} Demand (kg)'] = dem
        base_dict[f'Stop {i+1} Split Delivery?'] = split_flag
        
    master_data.append(base_dict)

# Standardize columns
cols = ['Route Type', 'Assigned Vehicle', 'Vehicle Capacity (kg)', 'Total Route Demand (kg)', 'Utilization (%)', 'Total Distance (km)']
for i in range(max_stops):
    cols.append(f'Stop {i+1} DH')
    cols.append(f'Stop {i+1} Restriction')
    cols.append(f'Stop {i+1} Demand (kg)')
    cols.append(f'Stop {i+1} Split Delivery?')

df = pd.DataFrame(master_data)
cols_exist = [c for c in cols if c in df.columns]
df = df[cols_exist]

# Excel Formatting with openpyxl
wb = Workbook()
ws = wb.active
ws.title = "Master Routing Plan"

# Write headers
for c_idx, col_name in enumerate(df.columns, 1):
    cell = ws.cell(row=1, column=c_idx, value=col_name)
    # Deep Zepto purple/black background, white bold text
    cell.fill = PatternFill(start_color="200E3A", end_color="200E3A", fill_type="solid")
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Write data
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Color specific columns lightly
        if c_idx == 1: # Route Type
            if value == "Direct":
                cell.fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# Auto-adjust column widths
for column_cells in ws.columns:
    length = max(len(str(cell.value)) for cell in column_cells if cell.value is not None)
    ws.column_dimensions[column_cells[0].column_letter].width = min(length + 4, 25)
    
ws.row_dimensions[1].height = 30

wb.save('Zepto_Final_Master_Plan.xlsx')
print("Zepto_Final_Master_Plan.xlsx generated successfully with perfect formatting!")

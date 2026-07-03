import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from constants import (
    TRUCK_CAPACITIES, CONTRACT_CONFIG,
    resolve_truck_type
)

DEFAULT_CONTRACT_TYPE = '24H'


def generate_report():
    """
    Read optimized_routes.csv and produce a multi-sheet Excel report:
      • Master Routing Plan        — one row per trip
      • Vehicle Summary            — one row per physical vehicle
      • Contract Optimization Plan — contract recommendations
    """
    print("Generating Zepto Final Master Plan...")

    # ─── Load Data ────────────────────────────────────────────────────
    routes = pd.read_csv('../data/outputs/optimized_routes.csv')
    sop = pd.read_excel('../data/inputs/details.xlsx', sheet_name='S&OP')

    sop['MH_clean'] = sop['MH '].astype(str).str.strip()
    sop['DH_clean'] = sop['DH'].astype(str).str.strip()
    demand_lookup = dict(
        zip(zip(sop['MH_clean'], sop['DH_clean']), sop['w1'])
    )
    restriction_lookup = dict(
        zip(zip(sop['MH_clean'], sop['DH_clean']),
            sop['Vehicle restrictions'])
    )

    # ─── Build Master Data ────────────────────────────────────────────
    master_data = []
    max_stops = 0

    for _, row in routes.iterrows():
        route_str = str(row.get('Route', ''))
        if route_str.strip() in ('', '—'):
            continue

        stops = [s.strip() for s in route_str.split('->')]
        mh_name = row.get('MH_Name', '')

        dhs = [s for s in stops if s != mh_name and s != '']
        max_stops = max(max_stops, len(dhs))
        
        is_split = row.get('Is_Split_Delivery', False)

        base_dict = {
            'Material Hub': mh_name,
            'Physical Vehicle ID': row.get('Physical_Vehicle_ID', ''),
            'Contract': row.get('Contract_Type', DEFAULT_CONTRACT_TYPE),
            'Trip': int(row.get('Trip_Number', 1)),
            'Route Type': row.get('Route_Type', 'Direct'),
            'Assigned Vehicle': row.get('Assigned_Truck', ''),
            'Vehicle Capacity (kg)': row.get('Truck_Capacity', 0),
            'Total Load (kg)': row.get('Total_Load', 0),
            'Weight Util (%)': row.get('Weight_Utilization_%', 0.0),
            'Total Distance (km)': row.get('Total_Distance_km', 0.0),
            'MH Overhead (min)': row.get('MH_Overhead_min', 0),
            'Loading (min)': row.get('Loading_Time_min', 0),
            'Travel (min)': row.get('Travel_Time_min', 0.0),
            'DH Overhead (min)': row.get('DH_Overhead_min', 0),
            'Unloading (min)': row.get('Unloading_Time_min', 0),
            'Trip Time (min)': row.get('Total_Trip_Time_min', 0.0),
            'Time Budget (min)': row.get('Time_Budget_min', 0),
            'Time Util (%)': row.get('Time_Utilization_%', 0.0),
            'Split Delivery': 'YES' if is_split else 'NO',
        }

        # Add per-stop columns
        for i, dh in enumerate(dhs):
            clean_name = dh.split('_part')[0]
            dem = demand_lookup.get((mh_name, clean_name), 0)
            restr_raw = restriction_lookup.get((mh_name, clean_name), '')
            resolved = resolve_truck_type(restr_raw)
            restr_display = resolved if resolved else str(restr_raw).strip()
            split_flag = "Yes" if '_part' in dh else "No"

            base_dict[f'Stop {i + 1} DH'] = clean_name
            base_dict[f'Stop {i + 1} Restriction'] = restr_display
            base_dict[f'Stop {i + 1} Demand (kg)'] = dem
            base_dict[f'Stop {i + 1} Split?'] = split_flag

        master_data.append(base_dict)

    # Standardize columns
    fixed_cols = [
        'Material Hub', 'Physical Vehicle ID', 'Contract', 'Trip',
        'Route Type', 'Assigned Vehicle', 'Vehicle Capacity (kg)',
        'Total Load (kg)', 'Weight Util (%)',
        'Total Distance (km)',
        'MH Overhead (min)', 'Loading (min)', 'Travel (min)',
        'DH Overhead (min)', 'Unloading (min)',
        'Trip Time (min)', 'Time Budget (min)', 'Time Util (%)',
        'Split Delivery'
    ]
    stop_cols = []
    for i in range(max_stops):
        stop_cols.extend([
            f'Stop {i + 1} DH',
            f'Stop {i + 1} Restriction',
            f'Stop {i + 1} Demand (kg)',
            f'Stop {i + 1} Split?',
        ])

    all_cols = fixed_cols + stop_cols
    df = pd.DataFrame(master_data)
    cols_exist = [c for c in all_cols if c in df.columns]
    df = df[cols_exist]
    df = df.astype(object)
    df.fillna('—', inplace=True)

    # ─── Build Vehicle Summary ────────────────────────────────────────
    summary_data = []
    if 'Physical Vehicle ID' in df.columns:
        for phys_id, grp in df.groupby('Physical Vehicle ID'):
            trip1 = grp[grp['Trip'] == 1].iloc[0] if len(grp[grp['Trip'] == 1]) > 0 else None
            trip2 = grp[grp['Trip'] == 2].iloc[0] if len(grp[grp['Trip'] == 2]) > 0 else None

            base = trip1 if trip1 is not None else trip2
            contract = str(base.get('Contract', DEFAULT_CONTRACT_TYPE))
            contract_cfg = CONTRACT_CONFIG.get(contract, CONTRACT_CONFIG['24H'])
            total_budget = contract_cfg['effective_minutes']

            t1_load = int(trip1['Total Load (kg)']) if trip1 is not None else 0
            t1_time = float(trip1['Trip Time (min)']) if trip1 is not None else 0.0
            t1_stops = 0
            for i in range(1, max_stops + 1):
                if trip1 is not None and trip1.get(f'Stop {i} DH') != '—':
                    t1_stops += 1
            
            t2_load = int(trip2['Total Load (kg)']) if trip2 is not None else 0
            t2_time = float(trip2['Trip Time (min)']) if trip2 is not None else 0.0
            t2_stops = 0
            for i in range(1, max_stops + 1):
                if trip2 is not None and trip2.get(f'Stop {i} DH') != '—':
                    t2_stops += 1

            total_load = t1_load + t2_load
            total_time = t1_time + t2_time
            combined_util = (
                (total_time / total_budget * 100)
                if total_budget > 0 else 0
            )

            summary_data.append({
                'Material Hub': base.get('Material Hub', ''),
                'Physical Vehicle ID': phys_id,
                'Contract': contract,
                'Truck': base.get('Assigned Vehicle', ''),
                'Capacity (kg)': base.get('Vehicle Capacity (kg)', 0),
                'Trip 1 Load (kg)': t1_load,
                'Trip 1 Time (min)': round(t1_time, 1),
                'Trip 1 Route': str(trip1.get('Route Type', '—')) if trip1 is not None else '—',
                'Trip 1 Stops': t1_stops,
                'Trip 2 Load (kg)': t2_load,
                'Trip 2 Time (min)': round(t2_time, 1),
                'Trip 2 Route': str(trip2.get('Route Type', '—')) if trip2 is not None else '—',
                'Trip 2 Stops': t2_stops,
                'Total Load (kg)': total_load,
                'Total Time (min)': round(total_time, 1),
                'Time Budget (min)': total_budget,
                'Combined Time Util (%)': round(combined_util, 1),
            })

    summary_df = pd.DataFrame(summary_data)
    
    # ─── Build Contract Optimization Report ───────────────────────────
    contract_opt_data = []
    
    # Extract candidate Direct trips to absorb
    direct_trips = [s for s in summary_data if s['Trip 1 Route'] == 'Direct' and s['Trip 2 Route'] == '—']
    
    for s in summary_data:
        phys_id = s['Physical Vehicle ID']
        curr_contract = s['Contract']
        util = s['Combined Time Util (%)']
        
        recommended = curr_contract
        reason = "Optimal / OK"
        
        if curr_contract == '24H':
            if util < 40.0:
                recommended = '12H'
                reason = 'Underutilized (< 40%)'
            elif 50.0 <= util <= 55.0:
                can_consolidate = False
                for dt in direct_trips:
                    if dt['Physical Vehicle ID'] != phys_id and dt['Material Hub'] == s['Material Hub']:
                        dt_load = dt['Trip 1 Load (kg)']
                        # Check Trip 1 absorption
                        if s['Trip 1 Route'] != '—' and s['Trip 1 Stops'] < 3:
                            if s['Trip 1 Load (kg)'] + dt_load <= s['Capacity (kg)']:
                                recommended = 'Consolidate'
                                reason = f"Can absorb stop from {dt['Physical Vehicle ID']}"
                                can_consolidate = True
                                break
                        # Check Trip 2 absorption
                        if s['Trip 2 Route'] != '—' and s['Trip 2 Stops'] < 3:
                            if s['Trip 2 Load (kg)'] + dt_load <= s['Capacity (kg)']:
                                recommended = 'Consolidate'
                                reason = f"Can absorb stop from {dt['Physical Vehicle ID']}"
                                can_consolidate = True
                                break
                if not can_consolidate:
                    reason = "Mid-utilized (No easy consolidation found)"
                    
        contract_opt_data.append({
            'Vehicle': phys_id,
            'Material Hub': s['Material Hub'],
            'Current Contract': curr_contract,
            'Recommended Contract': recommended,
            'Utilization %': util,
            'Reason': reason,
        })
        
    contract_df = pd.DataFrame(contract_opt_data)
    # drop columns from summary_df that were only needed for logic
    summary_df.drop(columns=['Trip 1 Stops', 'Trip 2 Stops'], inplace=True, errors='ignore')

    # ─── Write Excel ──────────────────────────────────────────────────
    wb = Workbook()

    # Shared styles
    header_fill = PatternFill(start_color="200E3A", end_color="200E3A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center")

    direct_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    milk_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    empty_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    time_warn_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    split_fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid") # Light Orange

    # ── Sheet 1: Master Routing Plan ──
    ws1 = wb.active
    ws1.title = "Master Routing Plan"
    _write_styled_sheet(
        ws1, df, header_fill, header_font, header_align,
        thin_border, center_align,
        row_style_fn=_master_row_style,
        style_fills={
            'direct': direct_fill, 'milk': milk_fill,
            'empty': empty_fill, 'time_warn': time_warn_fill,
            'split': split_fill
        }
    )

    # ── Sheet 2: Vehicle Summary ──
    ws2 = wb.create_sheet("Vehicle Summary")
    _write_styled_sheet(
        ws2, summary_df, header_fill, header_font, header_align,
        thin_border, center_align,
    )
    
    # ── Sheet 3: Contract Optimization Report ──
    ws3 = wb.create_sheet("Contract Optimization Report")
    _write_styled_sheet(
        ws3, contract_df, header_fill, header_font, header_align,
        thin_border, center_align,
    )

    # Auto-adjust column widths
    for ws in [ws1, ws2, ws3]:
        for column_cells in ws.columns:
            length = max(
                len(str(cell.value))
                for cell in column_cells if cell.value is not None
            )
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 4, 30)
        ws.row_dimensions[1].height = 32

    wb.save('../data/outputs/Zepto_Final_Master_Plan.xlsx')
    print("Zepto_Final_Master_Plan.xlsx generated successfully!")


def _write_styled_sheet(ws, df, header_fill, header_font, header_align,
                        thin_border, center_align,
                        row_style_fn=None, style_fills=None):
    if df.empty:
        return

    for c_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=c_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            cell.alignment = center_align

        if row_style_fn and style_fills:
            row_style_fn(ws, r_idx, row, df.columns, style_fills)


def _master_row_style(ws, r_idx, row, columns, fills):
    col_list = list(columns)

    if 'Route Type' in col_list:
        rt_idx = col_list.index('Route Type') + 1
        rt_val = str(row[rt_idx - 1]) if rt_idx - 1 < len(row) else ''
        cell = ws.cell(row=r_idx, column=rt_idx)
        if rt_val == 'Direct':
            cell.fill = fills['direct']
        elif rt_val == 'Milk Run':
            cell.fill = fills['milk']
        elif rt_val == '—':
            cell.fill = fills['empty']

    if 'Material Hub' in col_list:
        mh_idx = col_list.index('Material Hub') + 1
        ws.cell(row=r_idx, column=mh_idx).font = Font(bold=True)
        
    if 'Split Delivery' in col_list:
        split_idx = col_list.index('Split Delivery') + 1
        split_val = str(row[split_idx - 1]) if split_idx - 1 < len(row) else 'NO'
        if split_val == 'YES':
            ws.cell(row=r_idx, column=split_idx).fill = fills['split']
            # Highlight individual stop cells if they are part of a split
            for i in range(1, 4):
                split_flag_col = f'Stop {i} Split?'
                if split_flag_col in col_list:
                    flag_idx = col_list.index(split_flag_col) + 1
                    if row[flag_idx - 1] == 'Yes':
                        ws.cell(row=r_idx, column=flag_idx).fill = fills['split']
                        dh_idx = col_list.index(f'Stop {i} DH') + 1
                        ws.cell(row=r_idx, column=dh_idx).fill = fills['split']
                        dem_idx = col_list.index(f'Stop {i} Demand (kg)') + 1
                        ws.cell(row=r_idx, column=dem_idx).fill = fills['split']

    if 'Time Util (%)' in col_list:
        tu_idx = col_list.index('Time Util (%)') + 1
        tu_val = row[tu_idx - 1] if tu_idx - 1 < len(row) else 0
        try:
            if float(tu_val) > 90:
                ws.cell(row=r_idx, column=tu_idx).fill = fills['time_warn']
        except (ValueError, TypeError):
            pass

    if 'Total Load (kg)' in col_list:
        load_idx = col_list.index('Total Load (kg)') + 1
        load_val = row[load_idx - 1] if load_idx - 1 < len(row) else 0
        try:
            if int(load_val) == 0:
                for c_idx in range(1, len(row) + 1):
                    ws.cell(row=r_idx, column=c_idx).fill = fills['empty']
                    ws.cell(row=r_idx, column=c_idx).font = Font(color="999999", italic=True)
        except (ValueError, TypeError):
            pass


if __name__ == "__main__":
    generate_report()
